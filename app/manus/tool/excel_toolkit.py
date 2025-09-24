# Copyright 2025 ZTE Corporation.
# All Rights Reserved.
#
#    Licensed under the Apache License, Version 2.0 (the "License"); you may
#    not use this file except in compliance with the License. You may obtain
#    a copy of the License at
#
#         http://www.apache.org/licenses/LICENSE-2.0
#
#    Unless required by applicable law or agreed to in writing, software
#    distributed under the License is distributed on an "AS IS" BASIS, WITHOUT
#    WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the
#    License for the specific language governing permissions and limitations
#    under the License.


import os

from camel.toolkits import (
    ExcelToolkit
)

from app.manus.gate.format_gate import format_check
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook


@format_check()
def extract_excel_content(document_path: str) -> str:
    """
    Read the first sheet with colored cells in an Excel file and extract its color grid.

    Args:
        document_path (str): The path to the Excel file to be processed.

    Returns:
        str: A string representing the 2D color grid. Each row is a new line.
             Returns an empty string if no colored cells are found or an error occurs.
             Example return value: "['FF0000', '']\n['', '0070C0']"
    """
    try:
        workbook = load_workbook(document_path, data_only=True)
    except FileNotFoundError:
        print(f"Error: File not found at {document_path}")
        return ""
    except Exception as e:
        print(f"An error occurred while processing the file: {e}")
        return ""

    final_grid = []

    # --- Step 1: Find the first sheet with data cells and process it ---
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        sheet_values_map = {}

        # 1a. Map all non-empty cell coordinates to their values for the current sheet.
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    sheet_values_map[cell.coordinate] = cell.value

        # 1b. If data cells were found, process this sheet and then stop.
        if sheet_values_map:
            # --- Step 2: Calculate the boundaries of the data area ---
            coords = sheet_values_map.keys()
            row_nums = {int("".join(filter(str.isdigit, c))) for c in coords}
            col_letters = {"".join(filter(str.isalpha, c)) for c in coords}
            col_nums = {column_index_from_string(c) for c in col_letters}

            min_row, max_row = min(row_nums), max(row_nums)
            min_col, max_col = min(col_nums), max(col_nums)

            # --- Step 3: Build the 2D list (grid) for the values ---
            for row_index in range(min_row, max_row + 1):
                row_list = []
                for col_index in range(min_col, max_col + 1):
                    coordinate = f"{get_column_letter(col_index)}{row_index}"
                    cell_value = sheet_values_map.get(coordinate, '')
                    # Convert value to string to match original format
                    row_list.append(str(cell_value) if cell_value is not None else '')
                final_grid.append(row_list)

            # Since we only need the first sheet, break the loop after processing.
            break

    # --- Step 4: Convert the 2D list to the required string format ---
    if not final_grid:
        return ""

    # Join each row's string representation with a newline character.
    output_string = '\n'.join([str(row) for row in final_grid])
    print(output_string)
    return output_string


@format_check()
def read_excel_color(document_path):
    """
    Read an Excel file and extract cell colors/values

    Args:
        document_path (str): The path to the Excel file to be processed.

    Returns:
        dict: A dictionary mapping sheet names to their 2D color grid.
              Example: {'Sheet1': [['FF0000', '0070C0'], ['FF0000', '']]}.
              Returns None if the file is not found or contains no colored cells.
    """
    try:
        workbook = openpyxl.load_workbook(document_path, data_only=True)
    except FileNotFoundError:
        # A print statement can be useful for debugging in a console application.
        print(f"Error: File not found at {document_path}")
        return ""
    except Exception as e:
        print(f"An error occurred while processing the file: {e}")
        return ""

    final_grid = []

    # --- Step 1: Find the first sheet with colored cells and process it ---
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        sheet_colors_map = {}

        # 1a. Map all colored cell coordinates to their color codes for the current sheet.
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    color_code = cell.fill.fgColor.rgb
                    # Ignore transparent and pure white fills.
                    if color_code not in ('00000000', 'FFFFFFFF'):
                        sheet_colors_map[cell.coordinate] = color_code

        # 1b. If colored cells were found, process this sheet and then stop.
        if sheet_colors_map:
            # --- Step 2: Calculate the boundaries of the colored area ---
            coords = sheet_colors_map.keys()
            row_nums = {int("".join(filter(str.isdigit, c))) for c in coords}
            col_letters = {"".join(filter(str.isalpha, c)) for c in coords}
            col_nums = {column_index_from_string(c) for c in col_letters}

            min_row, max_row = min(row_nums), max(row_nums)
            min_col, max_col = min(col_nums), max(col_nums)

            # --- Step 3: Build the 2D list (grid) for the colors ---
            for row_index in range(min_row, max_row + 1):
                row_list = []
                for col_index in range(min_col, max_col + 1):
                    coordinate = f"{get_column_letter(col_index)}{row_index}"
                    color_hex = sheet_colors_map.get(coordinate)
                    # Get the 6-digit RGB value or an empty string for non-colored cells.
                    rgb_value = color_hex[2:] if color_hex else ''
                    row_list.append(rgb_value)
                final_grid.append(row_list)

            # Since we only need the first sheet, break the loop after processing.
            break

    # --- Step 4: Convert the 2D list to the required string format ---
    if not final_grid:
        return ""

    # Join each row's string representation with a newline character.
    output_string = '\n'.join([str(row) for row in final_grid])
    print(output_string)
    return output_string




if __name__ == '__main__':
    result = extract_excel_content("/home/eval_test.xlsx")
    print(result)
